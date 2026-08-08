#!/usr/bin/env python3
"""Audita PBT Web 2.0 contra los tres Excel originales sin modificar el motor."""
from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


MONTH_TO_QUARTER = {"Ene": "Q1_26", "Feb": "Q1_26", "Abr": "Q2_26"}
CATEGORY_INDEX = {
    "Espresso": 1,
    "Café filtrado": 2,
    "CBS": 3,
    "Food": 4,
    "Otro": 5,
    "Leche": 5,
}
ORDER_INDEX = {"Lobby": 6, "Pick up & Delivery": 7, "Drive Thru": 8}
PLAY_FIELDS = ("ptn", "cafe", "mop", "dt", "esp", "brew", "cbs", "food", "oth")
EXPECTED_WIDTHS = [9.08203125] * 13 + [10.83203125]


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_js_const(path: Path, name: str) -> Any:
    text = path.read_text(encoding="utf-8")
    match = re.search(rf"\bconst\s+{re.escape(name)}\s*=\s*", text)
    if not match:
        raise ValueError(f"No se encontró la constante {name} en {path}")
    start = match.end()
    return json.JSONDecoder().raw_decode(text, start)[0]


def compare_counts(source: Iterable[tuple], generated: Iterable[tuple]) -> dict[str, int]:
    source_counter, generated_counter = Counter(source), Counter(generated)
    return {
        "source": sum(source_counter.values()),
        "generated": sum(generated_counter.values()),
        "missing": sum((source_counter - generated_counter).values()),
        "extra": sum((generated_counter - source_counter).values()),
    }


def compact_play_key(lookup: str) -> str:
    parts = dict(part.split(":", 1) for part in lookup.split(";") if ":" in part)
    values = [int(float(parts["ptn"]))]
    values.extend(int(round(float(parts[field]) * 100)) for field in PLAY_FIELDS[1:])
    return ",".join(map(str, values))


def audit_pbt_motor(pbt_path: Path, excel_data_path: Path, index_path: Path) -> dict[str, Any]:
    names = (
        "cafeCandidates", "dtCandidates", "cafePlayEntries", "dtPlayEntries",
        "routines", "coords", "channelProd",
    )
    excel_data = load_js_const(excel_data_path, "EXCEL_DATA")
    generated = {name: excel_data[name] for name in names}
    wb = load_workbook(pbt_path, read_only=True, data_only=True)

    source_coords = []
    for lookup, partner in wb["coordinates"].iter_rows(min_row=2, values_only=True):
        play, layout, row, col = str(lookup).rsplit(":", 3)
        source_coords.append((play, layout.lstrip("_"), int(row), int(col), str(partner or "")))
    generated_coords = [
        (play, str(item[0]).lstrip("_"), int(item[1]), int(item[2]), str(item[3] or ""))
        for play, items in generated["coords"].items() for item in items
    ]

    source_routines = [
        (str(play), str(partner), str(station or ""), str(planted or ""), str(routine or ""))
        for _, play, partner, station, routine, planted, _
        in wb["routines"].iter_rows(min_row=2, values_only=True)
    ]
    generated_routines = [
        (str(play), str(row[0]), str(row[1] or ""), str(row[2] or ""), str(row[3] or ""))
        for play, rows in generated["routines"].items() for row in rows
    ]

    play_checks = {}
    for sheet_name, generated_name in (("cafe_plays", "cafePlayEntries"), ("dt_plays", "dtPlayEntries")):
        source = [
            (compact_play_key(str(lookup)), str(play))
            for lookup, play in wb[sheet_name].iter_rows(min_row=2, values_only=True)
        ]
        compact = [(str(key), str(play)) for key, play in generated[generated_name]]
        play_checks[sheet_name] = compare_counts(source, compact)

    candidate_checks = {}
    for sheet_name, generated_name in (("cafe_mixes", "cafeCandidates"), ("dt_mixes", "dtCandidates")):
        source = [
            tuple(int(round(float(value or 0) * 100)) for value in row[:8])
            for row in wb[sheet_name].iter_rows(min_row=2, values_only=True)
        ]
        compact = [tuple(map(int, row)) for row in generated[generated_name]]
        candidate_checks[sheet_name] = compare_counts(source, compact)

    source_channel = {
        str(play): int(value or 0)
        for play, value in wb["channel_production"].iter_rows(min_row=2, values_only=True)
    }
    generated_channel = {str(play): int(value or 0) for play, value in generated["channelProd"].items()}
    channel_check = compare_counts(source_channel.items(), generated_channel.items())
    wb.close()

    checks = {
        "coordinates": compare_counts(source_coords, generated_coords),
        "routines": compare_counts(source_routines, generated_routines),
        "plays": play_checks,
        "mixCandidates": candidate_checks,
        "channelProduction": channel_check,
    }
    index_text = index_path.read_text(encoding="utf-8")
    widths_match = re.search(r"widths=\[([^\]]+)\]", index_text)
    rendered_widths = [float(value) for value in widths_match.group(1).split(",")] if widths_match else []
    row_offset_ok = "Number(x[1])+17" in index_text
    column_offset_ok = "Number(x[2])+1" in index_text
    transform_mismatches = int(rendered_widths != EXPECTED_WIDTHS) + int(not row_offset_ok) + int(not column_offset_ok)
    checks["coordinates"]["rowOffset"] = 17
    checks["coordinates"]["columnOffset"] = 1
    checks["coordinates"]["visualStartCell"] = "B18"
    checks["coordinates"]["visualColumnWidths"] = rendered_widths
    checks["coordinates"]["transformMismatches"] = transform_mismatches
    return checks


def audit_sales_cube(source_path: Path, data_js_path: Path) -> dict[str, Any]:
    names = ["PBT_QS", "PBT_DAYPARTS", "PBT_MESES", "PBT_SEMANAS", "PBT_TIENDAS", "PBT_WEEKPARTS", "PBT_DATA"]
    values = {name: load_js_const(data_js_path, name) for name in names}
    dimensions = [values[name] for name in names[:6]]
    generated = {
        tuple(dimensions[index][int(row[index])] for index in range(6)): [float(value or 0) for value in row[6:15]]
        for row in values["PBT_DATA"]
    }

    source = defaultdict(lambda: [0.0] * 9)
    rows = 0
    wb = load_workbook(source_path, read_only=True, data_only=True)
    for category, _, daypart, month, week, store, weekpart, order, sale in wb["Base"].iter_rows(min_row=2, values_only=True):
        rows += 1
        month = str(month)
        key = (MONTH_TO_QUARTER[month], str(daypart), month, str(week), str(store), str(weekpart))
        amount = float(sale or 0)
        target = source[key]
        target[0] += amount
        if category in CATEGORY_INDEX:
            target[CATEGORY_INDEX[category]] += amount
        if order in ORDER_INDEX:
            target[ORDER_INDEX[order]] += amount
    wb.close()

    missing, mismatched = 0, 0
    max_difference = 0.0
    for key, raw in source.items():
        compact = generated.get(key)
        if compact is None:
            missing += 1
            continue
        expected = [round(value, 2) for value in raw]
        differences = [abs(compact[index] - expected[index]) for index in range(9)]
        max_difference = max(max_difference, *differences)
        if any(difference > 0.011 for difference in differences):
            mismatched += 1
    extra = sum(1 for key in generated if key[2] in MONTH_TO_QUARTER and key not in source)
    return {
        "sourceRows": rows,
        "sourceGroups": len(source),
        "generatedGroupsForSourceMonths": sum(1 for key in generated if key[2] in MONTH_TO_QUARTER),
        "missingGroups": missing,
        "extraGroups": extra,
        "mismatchedGroups": mismatched,
        "maxAbsoluteDifference": round(max_difference, 6),
        "roundingTolerance": 0.01,
        "categoryRule": "Leche se integra en Otro, igual que el cubo operativo original.",
    }


def audit_formula_map(formula_path: Path, pbt_path: Path) -> dict[str, Any]:
    guide = load_workbook(formula_path, read_only=True, data_only=False)
    motor = load_workbook(pbt_path, read_only=True, data_only=False)
    result = {}
    for guide_sheet, motor_sheet in (
        ("Cafe Layout_FormulasCelda", "Cafe Layout"),
        ("Drive Thru Layout_FormulasCelda", "Drive Thru Layout"),
    ):
        mapped, unique, invalid, guide_formulas = 0, set(), 0, 0
        target = motor[motor_sheet]
        for row in guide[guide_sheet].iter_rows(values_only=True):
            coordinate = row[1] if len(row) > 1 else None
            note = row[2] if len(row) > 2 else None
            if not isinstance(coordinate, str) or not re.fullmatch(r"[A-Z]+\d+", coordinate):
                continue
            mapped += 1
            unique.add(coordinate)
            if isinstance(note, str) and note.lstrip("_").startswith("="):
                guide_formulas += 1
            cell = target[coordinate]
            if cell.row > target.max_row or cell.column > target.max_column:
                invalid += 1
        result[motor_sheet] = {
            "mappedReferences": mapped,
            "uniqueReferences": len(unique),
            "formulaReferences": guide_formulas,
            "invalidReferences": invalid,
            "sourceDimensions": f"{target.max_row}x{target.max_column}",
        }
    guide.close()
    motor.close()
    return result


def mismatch_total(node: Any) -> int:
    if isinstance(node, dict):
        mismatch_keys = {"missing", "extra", "missingGroups", "extraGroups", "mismatchedGroups", "invalidReferences", "transformMismatches"}
        total = sum(int(node.get(key, 0)) for key in mismatch_keys)
        return total + sum(mismatch_total(value) for key, value in node.items() if key not in mismatch_keys)
    if isinstance(node, list):
        return sum(mismatch_total(value) for value in node)
    return 0


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--project", type=Path, default=Path(__file__).resolve().parents[1])
    parser.add_argument("--formula", type=Path, required=True)
    parser.add_argument("--sales", type=Path, required=True)
    parser.add_argument("--pbt", type=Path, required=True)
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()
    output = args.output or args.project / "reports" / "audit-originals.json"

    report = {
        "audit": "PBT Web 2.0 · fidelidad al motor original",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "sources": {
            "formulaMap": {"file": args.formula.name, "sha256": sha256(args.formula)},
            "salesBase": {"file": args.sales.name, "sha256": sha256(args.sales)},
            "originalMotor": {"file": args.pbt.name, "sha256": sha256(args.pbt)},
        },
        "engineArtifacts": {
            "data.js": sha256(args.project / "js" / "data.js"),
            "excel_data.js": sha256(args.project / "js" / "excel_data.js"),
        },
        "checks": {
            "motor": audit_pbt_motor(args.pbt, args.project / "js" / "excel_data.js", args.project / "index.html"),
            "salesCube": audit_sales_cube(args.sales, args.project / "js" / "data.js"),
            "formulaMap": audit_formula_map(args.formula, args.pbt),
        },
    }
    report["summary"] = {
        "passed": mismatch_total(report["checks"]) == 0,
        "mismatches": mismatch_total(report["checks"]),
        "statement": "El diseño puede cambiar; coordenadas, plays, rutinas y cálculos permanecen fieles a los Excel originales.",
    }
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(report["summary"], ensure_ascii=False, indent=2))
    if not report["summary"]["passed"]:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
